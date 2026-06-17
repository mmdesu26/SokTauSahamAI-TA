# Tujuan script:
# 1. Mengambil data historis saham dari yFinance.
# 2. Melatih model prediksi harga saham sesuai tanggal dasar (base-date).
# 3. Mengambil data internal model:
#    - fitur input model
#    - nilai mean dan scale dari StandardScaler
#    - koefisien dan intercept Linear Regression
#    - prediksi setiap tree Random Forest
#    - bobot ensemble RF dan LR
# 4. Menghitung ulang proses machine learning secara manual di Excel.
# 5. Membandingkan hasil manual Excel dengan hasil sistem.

# Jadi script ini dipakai untuk kebutuhan laporan/pembuktian:
# "hasil hitung manual ML sama dengan hasil sistem".

# argparse dipakai untuk membaca input dari terminal,
# contohnya --ticker, --base-date, --days, --lag-days, dan --output.
import argparse

# json dipakai untuk menampilkan hasil akhir dalam format JSON di terminal.
import json

# sys dipakai untuk mengatur exit program jika terjadi error.
import sys

# Path dipakai untuk mengatur lokasi file output Excel.
from pathlib import Path

# numpy dipakai untuk operasi angka, terutama mengecek arah naik/turun.
import numpy as np

# pandas dipakai untuk mengolah data historis saham dalam bentuk tabel.
import pandas as pd

# yfinance dipakai untuk mengambil data saham dari Yahoo Finance.
import yfinance as yf

# openpyxl dipakai untuk membuat dan mengatur file Excel.
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# StockPricePredictor adalah class prediksi harga saham dari sistem kamu.
# Class ini yang menjalankan training model dan prediksi harga.
from app.utils.ml_predictor import StockPricePredictor

# YFinanceHelper adalah helper dari sistem kamu.
# Dipakai untuk normalisasi kode saham dan normalisasi tanggal ke timezone Jakarta.
from app.utils.yfinance_helper import YFinanceHelper


# Jumlah data historis yang diambil dari yFinance.
# 730 hari kira-kira sekitar 2 tahun kalender.
LOOKBACK_DAYS = 730


def classify_error(error_pct: float) -> str:
    """
    Fungsi ini mengubah nilai error percentage menjadi kategori sederhana.

    Aturan:
    - error <= 1%  : Sangat sesuai
    - error <= 3%  : Masih sesuai
    - error > 3%   : Kurang sesuai

    Fungsi ini dipakai untuk memberi label kualitas prediksi.
    """

    # Jika error maksimal 1%, hasil prediksi dianggap sangat dekat dengan aktual.
    if error_pct <= 1:
        return "Sangat sesuai"

    # Jika error lebih dari 1% tetapi masih maksimal 3%,
    # hasil prediksi masih dianggap cukup sesuai.
    if error_pct <= 3:
        return "Masih sesuai"

    # Jika error lebih dari 3%, hasil prediksi dianggap kurang sesuai.
    return "Kurang sesuai"


def load_history(symbol: str) -> pd.DataFrame:
    """
    Fungsi ini mengambil data historis saham dari yFinance.

    Parameter:
    - symbol: kode saham, misalnya BBCA.JK

    Output:
    - DataFrame berisi data historis saham yang sudah dibersihkan dan diurutkan.
    """

    # Membuat object saham dari yFinance berdasarkan kode saham.
    stock = yf.Ticker(symbol)

    # Mengambil data historis saham selama LOOKBACK_DAYS hari.
    # interval="1d" artinya data harian.
    # auto_adjust=False artinya harga tidak disesuaikan otomatis.
    hist = stock.history(
        period=f"{LOOKBACK_DAYS}d",
        interval="1d",
        auto_adjust=False
    )

    # Jika data kosong, proses dihentikan karena model tidak bisa dilatih tanpa data.
    if hist is None or hist.empty:
        raise ValueError(f"Data historis kosong untuk {symbol}")

    # Menghapus baris yang tidak memiliki harga Close.
    # Harga Close wajib ada karena prediksi sistem berbasis harga penutupan.
    hist = hist.dropna(subset=["Close"]).copy()

    # Menormalisasi index tanggal ke timezone Jakarta.
    # Ini penting supaya tanggal trading konsisten dengan konteks Bursa Efek Indonesia.
    hist = YFinanceHelper._normalize_history_index_to_jakarta(hist)

    # Jika setelah normalisasi data jadi kosong, proses dihentikan.
    if hist.empty:
        raise ValueError(f"Data historis kosong setelah normalisasi untuk {symbol}")

    # Mengurutkan data berdasarkan tanggal dari lama ke baru.
    return hist.sort_index()


def style_sheet(ws):
    """
    Fungsi ini hanya untuk merapikan tampilan sheet Excel.

    Yang diatur:
    - warna header
    - font header
    - border tabel
    - alignment cell
    - lebar kolom
    - freeze pane supaya header tetap terlihat
    """

    # Warna biru untuk header tabel.
    header_fill = PatternFill("solid", fgColor="1F4E78")

    # Font putih dan tebal untuk header.
    header_font = Font(color="FFFFFF", bold=True)

    # Border tipis berwarna biru muda.
    thin = Side(style="thin", color="D9E2F3")

    # Memberi alignment dan border ke semua cell yang ada di sheet.
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Mengatur baris pertama sebagai header.
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Mengatur lebar setiap kolom agar isi lebih mudah dibaca.
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22

    # Membekukan baris pertama agar header tetap terlihat saat scroll.
    ws.freeze_panes = "A2"


def write_key_value(ws, data):
    """
    Fungsi ini menulis dictionary ke Excel dalam bentuk 2 kolom:
    - Komponen
    - Nilai

    Dipakai untuk menulis output sistem ke sheet Excel.
    """

    # Header tabel.
    ws.append(["Komponen", "Nilai"])

    # Menulis setiap key dan value dari dictionary.
    for k, v in data.items():

        # Jika value berbentuk dictionary/list,
        # ubah dulu menjadi teks JSON supaya bisa masuk satu cell Excel.
        if isinstance(v, (dict, list)):
            v = json.dumps(v, ensure_ascii=False)

        # Menulis komponen dan nilainya ke sheet.
        ws.append([k, v])


def main():
    """
    Fungsi utama script.

    Alur besar:
    1. Membaca argumen terminal.
    2. Mengambil data historis saham.
    3. Menentukan tanggal dasar dan target.
    4. Melatih model berdasarkan data sampai tanggal dasar.
    5. Mengambil parameter internal model.
    6. Membuat file Excel berisi perhitungan manual.
    """

    # Membuat parser untuk membaca input dari terminal.
    parser = argparse.ArgumentParser(
        description="Export perbandingan manual ML exact vs sistem ke Excel"
    )

    # Kode saham yang akan diuji, contoh BBCA atau BBCA.JK.
    parser.add_argument(
        "--ticker",
        required=True,
        help="Contoh: BBCA atau BBCA.JK"
    )

    # Tanggal dasar prediksi.
    # Model akan dilatih menggunakan data sampai tanggal ini.
    parser.add_argument(
        "--base-date",
        required=True,
        help="Tanggal dasar prediksi, format YYYY-MM-DD"
    )

    # Jumlah data historis yang dipakai untuk model.
    parser.add_argument("--days", type=int, default=730)

    # Jumlah lag fitur, misalnya 15 berarti memakai data 15 hari sebelumnya.
    parser.add_argument("--lag-days", type=int, default=15)

    # Nama file Excel output.
    parser.add_argument(
        "--output",
        default="manual_ml_exact_vs_sistem.xlsx"
    )

    # Mengambil semua argumen dari terminal.
    args = parser.parse_args()

    # Menormalisasi kode saham.
    # Contoh: BBCA bisa dinormalisasi menjadi BBCA.JK.
    symbol = YFinanceHelper.normalize_symbol(args.ticker)

    # Mengubah string tanggal base-date menjadi format date.
    requested_base_date = pd.Timestamp(args.base_date).date()

    # Mengambil data historis saham dari yFinance.
    hist = load_history(symbol)

    # Mengambil data yang tanggalnya <= base-date.
    # Data ini dipakai untuk menentukan tanggal dasar trading yang valid.
    eligible = hist[hist.index.date <= requested_base_date]

    # Jika tidak ada data sebelum/sama dengan base-date, proses dihentikan.
    if eligible.empty:
        raise ValueError(
            f"Tidak ada data trading pada/sebelum {requested_base_date} untuk {symbol}"
        )

    # Baris terakhir pada/sampai base-date dianggap sebagai data tanggal dasar.
    base_row = eligible.iloc[-1]

    # Index/tanggal trading yang benar-benar dipakai sebagai base date.
    base_idx = eligible.index[-1]

    # Posisi base date di data historis lengkap.
    base_pos = hist.index.get_loc(base_idx)

    # Jika base date sudah di baris terakhir data,
    # berarti tidak ada hari trading berikutnya untuk dibandingkan.
    if base_pos >= len(hist.index) - 1:
        raise ValueError("Tidak ada hari trading berikutnya setelah tanggal dasar.")

    # Tanggal trading berikutnya setelah base date.
    actual_idx = hist.index[base_pos + 1]

    # Baris data aktual pada hari trading berikutnya.
    actual_row = hist.iloc[base_pos + 1]

    # Data training hanya sampai base date.
    # Ini penting agar model tidak bocor memakai data masa depan.
    training_hist = hist[hist.index <= base_idx].copy()

    # Membuat object predictor menggunakan data historis sampai base date.
    predictor = StockPricePredictor(
        ticker=symbol,
        days=args.days,
        forecast_horizon=1,
        lag_days=args.lag_days,
        historical_df=training_hist,
    )

    # Melatih model prediksi harga.
    # Jika gagal, proses dihentikan.
    if not predictor.train_price_model():
        raise ValueError("Model gagal dilatih")

    # Menghasilkan prediksi untuk periode berikutnya.
    # Jika gagal, proses dihentikan.
    prediction = predictor.predict_next_period()
    if not prediction:
        raise ValueError("Model gagal menghasilkan prediksi")

    # Mengambil trainer dari predictor.
    # Trainer menyimpan model, scaler, fitur, dan parameter lain.
    trainer = predictor.trainer

    # Membuat artifact dari model yang baru dilatih.
    # Artifact berisi scaler, model LR, model RF, feature columns, dan bobot ensemble.
    artifact = trainer.build_artifact()

    # Mengambil daftar fitur yang dipakai model.
    feature_cols = artifact["feature_columns"]

    # ============================================================
    # BAGIAN PENTING:
    # Untuk perhitungan manual vs sistem, fitur harus diambil dari
    # dataset prediksi, bukan dataset training.
    #
    # Alasannya:
    # - Dataset training dipakai untuk belajar dan biasanya membuang baris terakhir
    #   karena target_return_future masih kosong.
    # - Dataset prediksi memakai baris terakhir yang tersedia sebagai input prediksi.
    #
    # Jadi agar Excel sama dengan sistem, input fiturnya harus dari dataset prediksi.
    # ============================================================

    # Menyiapkan dataset prediksi dan mengambil harga close terakhir yang valid.
    prediction_dataset, latest_completed_close = predictor.service._prepare_prediction_dataset()

    # Jika dataset prediksi kosong, tidak bisa membuat perhitungan manual.
    if prediction_dataset is None or prediction_dataset.empty:
        raise ValueError(
            "Dataset prediksi kosong; tidak bisa export perhitungan manual exact"
        )

    # Mengambil baris terakhir sebagai input prediksi model.
    latest_features_row = prediction_dataset.iloc[-1]

    # Mengambil nilai raw fitur sesuai urutan feature_cols.
    # Bentuknya dibuat 2D karena scaler dan model membutuhkan input seperti tabel.
    X_raw = latest_features_row[feature_cols].to_frame().T.values

    # Mengambil scaler dari artifact.
    scaler = artifact["scaler"]

    # Menormalisasi nilai raw fitur menggunakan scaler sistem.
    X_scaled = scaler.transform(X_raw)

    # Harga close pada tanggal dasar / harga terakhir yang dipakai model.
    current_price = float(latest_completed_close["close"])

    # Harga aktual pada tanggal target.
    actual_close = float(actual_row["Close"])

    # Baseline adalah asumsi harga target sama dengan harga close tanggal dasar.
    baseline_close = current_price

    # Mengambil model Linear Regression dari artifact.
    lr_model = artifact["lr_model"]

    # Mengambil model Random Forest dari artifact.
    rf_model = artifact["rf_model"]

    # Mengambil bobot ensemble dari artifact.
    # Bobot ini dipakai untuk menggabungkan hasil RF dan LR.
    weights = artifact["ensemble_weights"]

    # Menghitung return prediksi dari Linear Regression.
    lr_return_system = float(lr_model.predict(X_scaled)[0])

    # Menghitung return prediksi dari Random Forest.
    rf_return_system = float(rf_model.predict(X_scaled)[0])

    # Mengubah return LR menjadi harga prediksi LR.
    lr_price_system_exact = current_price * (1 + lr_return_system)

    # Mengubah return RF menjadi harga prediksi RF.
    rf_price_system_exact = current_price * (1 + rf_return_system)

    # Menghitung harga prediksi ensemble secara manual dari RF dan LR.
    ensemble_price_manual_exact = (
        (rf_price_system_exact * weights["rf"]) +
        (lr_price_system_exact * weights["lr"])
    )

    # Menghitung return ensemble secara manual dari return RF dan LR.
    # Variabel ini tidak ditulis ke Excel pada script ini, tapi bisa dipakai jika dibutuhkan.
    ensemble_return_manual_exact = (
        (rf_return_system * weights["rf"]) +
        (lr_return_system * weights["lr"])
    )

    # Mengambil harga prediksi akhir dari output sistem.
    predicted_close_system = float(prediction["predicted_close_next_day"])

    # Menghitung absolute error:
    # selisih absolut antara harga prediksi dan harga aktual.
    absolute_error = abs(predicted_close_system - actual_close)

    # Menghitung error percentage / APE untuk satu data prediksi.
    error_pct = (absolute_error / actual_close) * 100 if actual_close else 0.0

    # Menghitung error baseline.
    baseline_absolute_error = abs(baseline_close - actual_close)

    # Menghitung error percentage baseline.
    baseline_error_pct = (
        (baseline_absolute_error / actual_close) * 100
        if actual_close
        else 0.0
    )

    # Mengecek apakah arah prediksi benar.
    # Jika aktual naik dan prediksi juga naik, hasilnya True.
    # Jika aktual turun dan prediksi juga turun, hasilnya True.
    # Selain itu False.
    direction_correct = bool(
        np.sign(actual_close - current_price)
        == np.sign(predicted_close_system - current_price)
    )

    # Dictionary ini berisi output utama sistem.
    # Nanti ditulis ke sheet 01_Output_Sistem.
    system_output = {
        "ticker": symbol,
        "requested_base_date": str(requested_base_date),
        "actual_base_date_used": base_idx.strftime("%Y-%m-%d"),
        "base_close": round(current_price, 2),
        "target_date": actual_idx.strftime("%Y-%m-%d"),
        "actual_close_target": round(actual_close, 2),
        "predicted_close_target": round(predicted_close_system, 2),
        "rf_prediction_system_price": float(
            prediction.get("rf_prediction", round(rf_price_system_exact, 2))
        ),
        "lr_prediction_system_price": float(
            prediction.get("lr_prediction", round(lr_price_system_exact, 2))
        ),
        "baseline_close_target": round(baseline_close, 2),
        "absolute_error": round(absolute_error, 2),
        "error_percentage": round(error_pct, 4),
        "baseline_absolute_error": round(baseline_absolute_error, 2),
        "baseline_error_percentage": round(baseline_error_pct, 4),
        "classification": classify_error(error_pct),
        "model_beats_baseline": error_pct < baseline_error_pct,
        "direction_correct": direction_correct,
        "ensemble_weights": weights,
        "metrics": prediction.get("validation", {}),
    }

    # Mengambil prediksi return dari setiap tree Random Forest.
    # Jika RF punya 400 tree, maka list ini berisi 400 angka.
    tree_returns = [
        float(tree.predict(X_scaled)[0])
        for tree in rf_model.estimators_
    ]

    # Membuat workbook Excel baru.
    wb = Workbook()

    # Menghapus sheet default bawaan openpyxl.
    wb.remove(wb.active)

    # ============================================================
    # SHEET 01: OUTPUT SISTEM
    # ============================================================
    # Sheet ini berisi hasil prediksi dan evaluasi dari sistem.
    ws = wb.create_sheet("01_Output_Sistem")
    write_key_value(ws, system_output)
    style_sheet(ws)

    # ============================================================
    # SHEET 02: FITUR DAN NORMALISASI
    # ============================================================
    # Sheet ini membandingkan normalisasi manual Excel dengan scaler sistem.
    ws = wb.create_sheet("02_Fitur_Normalisasi")
    ws.append([
        "No",
        "Fitur",
        "Nilai Raw x",
        "Mean Scaler",
        "Scale/Std Scaler",
        "Manual z=(x-mean)/scale",
        "Z Sistem",
        "Selisih",
    ])

    # Menulis semua fitur yang dipakai model.
    for i, col in enumerate(feature_cols, start=1):
        r = i + 1
        ws.append([
            i,                         # nomor fitur
            col,                       # nama fitur
            float(X_raw[0, i - 1]),    # nilai asli fitur
            float(scaler.mean_[i - 1]),# mean dari StandardScaler
            float(scaler.scale_[i - 1]),# scale/std dari StandardScaler
            f"=(C{r}-D{r})/E{r}",      # rumus manual normalisasi
            float(X_scaled[0, i - 1]), # hasil normalisasi dari sistem
            f"=F{r}-G{r}",             # selisih manual dan sistem
        ])
    style_sheet(ws)

    # ============================================================
    # SHEET 03: LINEAR REGRESSION MANUAL
    # ============================================================
    # Sheet ini menghitung ulang prediksi Linear Regression secara manual.
    ws = wb.create_sheet("03_LR_Manual")
    ws.append(["No", "Fitur", "Z Manual", "Koefisien LR", "Kontribusi z*coef"])

    # Setiap fitur hasil normalisasi dikalikan dengan koefisien LR.
    for i, col in enumerate(feature_cols, start=1):
        r = i + 1
        ws.append([
            i,
            col,
            f"='02_Fitur_Normalisasi'!F{r}", # z manual dari sheet normalisasi
            float(lr_model.coef_[i - 1]),    # koefisien LR dari model sistem
            f"=C{r}*D{r}",                   # kontribusi fitur
        ])

    # Baris terakhir fitur.
    end_row = len(feature_cols) + 1

    # Menulis intercept LR.
    ws.append(["", "Intercept", "", float(lr_model.intercept_), ""])
    intercept_row = end_row + 1

    # Menghitung LR return manual:
    # intercept + jumlah semua kontribusi fitur.
    ws.append([
        "",
        "LR Return Manual",
        "",
        "",
        f"=D{intercept_row}+SUM(E2:E{end_row})",
    ])
    lr_return_row = intercept_row + 1

    # Menulis LR return dari sistem.
    ws.append(["", "LR Return Sistem", "", "", lr_return_system])

    # Selisih antara hasil manual dan hasil sistem.
    ws.append([
        "",
        "Selisih Return",
        "",
        "",
        f"=E{lr_return_row}-E{lr_return_row + 1}",
    ])

    # Menulis harga dasar.
    ws.append(["", "Current/Base Close", "", "", current_price])
    base_row_excel = lr_return_row + 3

    # Mengubah LR return manual menjadi harga.
    ws.append([
        "",
        "LR Price Manual",
        "",
        "",
        f"=E{base_row_excel}*(1+E{lr_return_row})",
    ])

    # Menulis harga prediksi LR dari sistem.
    ws.append([
        "",
        "LR Price Sistem",
        "",
        "",
        float(prediction.get("lr_prediction", round(lr_price_system_exact, 2))),
    ])
    style_sheet(ws)

    # ============================================================
    # SHEET 04: RANDOM FOREST MANUAL
    # ============================================================
    # Sheet ini menghitung ulang prediksi Random Forest.
    # Caranya dengan merata-ratakan prediksi dari seluruh tree.
    ws = wb.create_sheet("04_RF_Manual")
    ws.append(["Tree", "Tree Return", "Tree Price = base*(1+return)"])

    # Menulis prediksi return dari setiap tree Random Forest.
    for i, val in enumerate(tree_returns, start=1):
        r = i + 1
        ws.append([
            i,
            val,
            f"='03_LR_Manual'!E{base_row_excel}*(1+B{r})",
        ])

    # Baris terakhir tree.
    end_tree_row = len(tree_returns) + 1

    # Menghitung rata-rata return semua tree.
    ws.append(["RF Return Manual", f"=AVERAGE(B2:B{end_tree_row})", ""])
    rf_return_row = end_tree_row + 1

    # Menulis RF return dari sistem.
    ws.append(["RF Return Sistem", rf_return_system, ""])

    # Selisih return manual RF dan sistem.
    ws.append(["Selisih Return", f"=B{rf_return_row}-B{rf_return_row + 1}", ""])

    # Mengubah RF return manual menjadi harga.
    ws.append([
        "RF Price Manual",
        f"='03_LR_Manual'!E{base_row_excel}*(1+B{rf_return_row})",
        "",
    ])

    # Menulis RF price dari sistem.
    ws.append([
        "RF Price Sistem",
        float(prediction.get("rf_prediction", round(rf_price_system_exact, 2))),
        "",
    ])
    style_sheet(ws)

    # ============================================================
    # SHEET 05: ENSEMBLE MANUAL
    # ============================================================
    # Sheet ini menggabungkan hasil LR dan RF memakai bobot ensemble.
    ws = wb.create_sheet("05_Ensemble_Manual")
    ws.append(["Komponen", "Nilai", "Rumus/Keterangan"])

    # Data dan rumus yang ditampilkan pada sheet ensemble.
    rows = [
        [
            "Current/Base Close",
            current_price,
            "Sama dengan close tanggal dasar",
        ],
        [
            "LR Return Manual",
            "='03_LR_Manual'!E{}".format(lr_return_row),
            "Dari intercept + SUMPRODUCT",
        ],
        [
            "RF Return Manual",
            "='04_RF_Manual'!B{}".format(rf_return_row),
            "Rata-rata return 400 tree",
        ],
        [
            "Bobot LR",
            weights["lr"],
            "Dari sistem/artifact",
        ],
        [
            "Bobot RF",
            weights["rf"],
            "Dari sistem/artifact",
        ],
        [
            "LR Price Manual",
            "=B2*(1+B3)",
            "base_close*(1+LR_return)",
        ],
        [
            "RF Price Manual",
            "=B2*(1+B4)",
            "base_close*(1+RF_return)",
        ],
        [
            "Ensemble Price Manual",
            "=(B8*B6)+(B7*B5)",
            "RF_price*w_RF + LR_price*w_LR",
        ],
        [
            "Predicted Close Sistem",
            predicted_close_system,
            "Output sistem sudah round 2 desimal",
        ],
        [
            "Predicted Close Manual Rounded",
            "=ROUND(B9,2)",
            "Dibandingkan dengan sistem",
        ],
        [
            "Selisih Manual Rounded - Sistem",
            "=B11-B10",
            "Harus 0 atau mendekati 0",
        ],
    ]

    # Menulis semua baris ensemble ke Excel.
    for row in rows:
        ws.append(row)
    style_sheet(ws)

    # ============================================================
    # SHEET 06: EVALUASI MANUAL
    # ============================================================
    # Sheet ini menghitung error manual dan membandingkannya dengan sistem.
    ws = wb.create_sheet("06_Evaluasi_Manual")
    ws.append(["Komponen", "Manual Excel", "Sistem", "Selisih", "Rumus"])

    # Harga aktual target.
    ws.append([
        "Actual Close Target",
        actual_close,
        system_output["actual_close_target"],
        "=B2-C2",
        "Harga aktual hari bursa berikutnya",
    ])

    # Harga prediksi akhir.
    ws.append([
        "Predicted Close",
        "='05_Ensemble_Manual'!B11",
        system_output["predicted_close_target"],
        "=B3-C3",
        "Harga prediksi manual rounded vs sistem",
    ])

    # Absolute error.
    ws.append([
        "Absolute Error",
        "=ABS(B2-B3)",
        system_output["absolute_error"],
        "=B4-C4",
        "ABS(actual-predicted)",
    ])

    # Error percentage / APE.
    ws.append([
        "Error Percentage",
        "=(B4/B2)*100",
        system_output["error_percentage"],
        "=B5-C5",
        "absolute_error/actual*100",
    ])

    # Baseline close.
    ws.append([
        "Baseline Close",
        baseline_close,
        system_output["baseline_close_target"],
        "=B6-C6",
        "Baseline = close tanggal dasar",
    ])

    # Baseline absolute error.
    ws.append([
        "Baseline Absolute Error",
        "=ABS(B2-B6)",
        system_output["baseline_absolute_error"],
        "=B7-C7",
        "ABS(actual-baseline)",
    ])

    # Baseline error percentage.
    ws.append([
        "Baseline Error Percentage",
        "=(B7/B2)*100",
        system_output["baseline_error_percentage"],
        "=B8-C8",
        "baseline_error/actual*100",
    ])

    # Mengecek arah prediksi benar/salah.
    ws.append([
        "Direction Correct",
        direction_correct,
        system_output["direction_correct"],
        "=B9=C9",
        "SIGN(actual-base)=SIGN(predicted-base)",
    ])
    style_sheet(ws)

    # ============================================================
    # SHEET 07: RINGKASAN BANDING
    # ============================================================
    # Sheet ini adalah tabel utama untuk laporan.
    # Isinya membandingkan hasil manual Excel dengan hasil sistem.
    ws = wb.create_sheet("07_Ringkasan_Banding")
    ws.append([
        "Bagian ML",
        "Manual Excel",
        "Hasil Sistem",
        "Selisih",
        "Status",
        "Keterangan",
    ])

    # Membandingkan normalisasi manual dengan scaler.transform sistem.
    ws.append([
        "Normalisasi fitur",
        "=MAX(ABS('02_Fitur_Normalisasi'!H2:H49))",
        0,
        "=B2-C2",
        '=IF(ABS(D2)<0.000001,"SAMA","CEK")',
        "Membandingkan z manual vs z dari scaler.transform",
    ])

    # Membandingkan LR manual dengan lr_model.predict.
    ws.append([
        "LR Return",
        "='03_LR_Manual'!E{}".format(lr_return_row),
        lr_return_system,
        "=B3-C3",
        '=IF(ABS(D3)<0.000001,"SAMA","CEK")',
        "Intercept + SUM(z*coef) vs lr_model.predict",
    ])

    # Membandingkan harga LR manual dengan sistem.
    ws.append([
        "LR Price",
        "='05_Ensemble_Manual'!B7",
        round(lr_price_system_exact, 6),
        "=B4-C4",
        '=IF(ABS(D4)<0.01,"SAMA","CEK")',
        "base_close*(1+LR_return)",
    ])

    # Membandingkan RF manual dengan rf_model.predict.
    ws.append([
        "RF Return",
        "='04_RF_Manual'!B{}".format(rf_return_row),
        rf_return_system,
        "=B5-C5",
        '=IF(ABS(D5)<0.000001,"SAMA","CEK")',
        "Rata-rata 400 tree vs rf_model.predict",
    ])

    # Membandingkan harga RF manual dengan sistem.
    ws.append([
        "RF Price",
        "='05_Ensemble_Manual'!B8",
        round(rf_price_system_exact, 6),
        "=B6-C6",
        '=IF(ABS(D6)<0.01,"SAMA","CEK")',
        "base_close*(1+RF_return)",
    ])

    # Membandingkan harga ensemble manual dengan output sistem.
    ws.append([
        "Ensemble Price",
        "='05_Ensemble_Manual'!B11",
        system_output["predicted_close_target"],
        "=B7-C7",
        '=IF(ABS(D7)<0.01,"SAMA","CEK")',
        "RF*weight_RF + LR*weight_LR, dibulatkan 2 desimal",
    ])

    # Membandingkan absolute error manual dengan sistem.
    ws.append([
        "Absolute Error",
        "='06_Evaluasi_Manual'!B4",
        system_output["absolute_error"],
        "=B8-C8",
        '=IF(ABS(D8)<0.01,"SAMA","CEK")',
        "ABS(actual-predicted)",
    ])

    # Membandingkan error percentage manual dengan sistem.
    ws.append([
        "Error Percentage",
        "='06_Evaluasi_Manual'!B5",
        system_output["error_percentage"],
        "=B9-C9",
        '=IF(ABS(D9)<0.0001,"SAMA","CEK")',
        "absolute_error/actual*100",
    ])

    # Membandingkan arah prediksi manual dengan sistem.
    ws.append([
        "Direction Correct",
        "='06_Evaluasi_Manual'!B9",
        system_output["direction_correct"],
        "=B10=C10",
        '=IF(D10,"SAMA","CEK")',
        "Arah aktual vs arah prediksi",
    ])
    style_sheet(ws)

    # ============================================================
    # SHEET 08: KEGUNAAN
    # ============================================================
    # Sheet ini menjelaskan fungsi setiap sheet.
    ws = wb.create_sheet("08_Kegunaan")
    ws.append(["Sheet/Bagian", "Kegunaan", "Dipakai untuk menjawab"])

    rows_info = [
        [
            "01_Output_Sistem",
            "Menyimpan output backtesting sistem pada ticker dan tanggal yang sama.",
            "Bukti hasil sistem",
        ],
        [
            "02_Fitur_Normalisasi",
            "Menunjukkan 48 fitur input ML, mean/scale StandardScaler, z manual, z sistem, dan selisih.",
            "Apakah input manual sama dengan input sistem",
        ],
        [
            "03_LR_Manual",
            "Menghitung prediksi Linear Regression dari intercept + jumlah kontribusi fitur.",
            "Perhitungan manual LR vs lr_model.predict",
        ],
        [
            "04_RF_Manual",
            "Menghitung Random Forest dari rata-rata prediksi seluruh tree.",
            "Perhitungan manual RF vs rf_model.predict",
        ],
        [
            "05_Ensemble_Manual",
            "Menggabungkan harga prediksi RF dan LR memakai bobot sistem.",
            "Prediksi akhir manual vs sistem",
        ],
        [
            "06_Evaluasi_Manual",
            "Menghitung absolute error, error percentage, baseline, dan arah prediksi.",
            "Evaluasi manual vs output backtesting",
        ],
        [
            "07_Ringkasan_Banding",
            "Tabel utama perbandingan manual Excel dan sistem secara numerik.",
            "Tabel yang dimasukkan ke laporan",
        ],
    ]

    # Menulis keterangan kegunaan sheet.
    for row in rows_info:
        ws.append(row)
    style_sheet(ws)

    # Menentukan nama file output dari argumen terminal.
    output = Path(args.output)

    # Menyimpan workbook Excel.
    wb.save(output)

    # Menampilkan pesan sukses di terminal dalam format JSON.
    print(json.dumps(
        {
            "success": True,
            "output": str(output.resolve()),
            "system_output": system_output,
        },
        indent=2,
        ensure_ascii=False,
    ))


# Bagian ini memastikan main() hanya berjalan ketika file ini dijalankan langsung.
# Jika file di-import oleh file lain, main() tidak otomatis berjalan.
if __name__ == "__main__":
    try:
        # Menjalankan fungsi utama.
        main()

    except Exception as exc:
        # Jika terjadi error, tampilkan pesan error dalam format JSON.
        print(
            json.dumps(
                {
                    "success": False,
                    "message": str(exc),
                },
                ensure_ascii=False,
            ),
            file=sys.stderr,
        )

        # Menghentikan program dengan status error.
        sys.exit(1)
